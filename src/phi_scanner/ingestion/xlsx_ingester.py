"""XLSX ingester — streams ``CellRecord`` objects, one per non-empty cell.

v4 improvements:
  - Yields ``CellRecord`` (not bare tuples) so that row_context (sibling cell
    values) flows into the pipeline for value-density profiling.
  - Row-level context is collected in a single pass over each row; memory use
    stays O(row_width) rather than O(sheet_size).
  - All prior v2 numeric/boolean/date normalizations preserved.
"""
from __future__ import annotations

import datetime
from pathlib import Path
from typing import Iterator

import openpyxl
from openpyxl.utils import get_column_letter

from .base import CellRecord, SourceLocation


class XlsxIngester:
    """Opens a workbook in read-only mode to cap memory on large files.

    ``data_only=True`` returns the cached formula result rather than the
    formula string — relevant when cells hold computed identifiers.
    Every sheet in the workbook is scanned in order.
    """

    @staticmethod
    def _cell_to_str(value: object) -> str | None:
        """Convert a cell value to a clean string, or None if it should be skipped."""
        if value is None:
            return None
        if isinstance(value, bool):
            return None  # True/False are not identifiers
        if isinstance(value, (datetime.datetime, datetime.date, datetime.time)):
            return None  # dates are not PII identifiers
        if isinstance(value, float) and value.is_integer():
            return str(int(value))
        text = str(value).strip()
        return text if text else None

    def ingest(self, path: Path) -> Iterator[tuple[str, SourceLocation]]:
        """Yield ``(cell_text, location)`` for every non-empty cell.

        Legacy tuple protocol for backward compatibility. Use ``ingest_records()``
        for full ``CellRecord`` objects with row_context.
        """
        yield from ((r.text, r.location) for r in self.ingest_records(path))

    def ingest_records(self, path: Path) -> Iterator[CellRecord]:
        """Yield full ``CellRecord`` objects with row-level sibling context."""
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        try:
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                for row in ws.iter_rows():
                    # Collect all non-None string values in this row (O(row_width))
                    row_texts: list[str] = []
                    cell_entries: list[tuple[str, int, int]] = []  # (text, row_num, col_num)

                    for cell in row:
                        text = self._cell_to_str(cell.value)
                        if text:
                            row_texts.append(text)
                            cell_entries.append((text, cell.row, cell.column))

                    if not cell_entries:
                        continue

                    row_context = " ".join(row_texts)

                    for text, row_num, col_num in cell_entries:
                        yield CellRecord(
                            text=text,
                            location=SourceLocation(
                                file_path=path,
                                sheet_name=sheet_name,
                                row=row_num,
                                column=get_column_letter(col_num),
                            ),
                            row_context=row_context,
                        )
        finally:
            wb.close()
