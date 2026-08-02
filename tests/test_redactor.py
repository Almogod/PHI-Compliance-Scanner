"""Tests for Redaction and Sanitization Engine."""
import csv
from pathlib import Path
import openpyxl

from phi_scanner.redactor import redact_csv, redact_xlsx, sanitize_text


def test_sanitize_text() -> None:
    raw = "Contact Ramesh at 9876543210 with Aadhaar 999988887777 and PAN ABCDE1234F"
    clean = sanitize_text(raw)
    assert "9876543210" not in clean
    assert "999988887777" not in clean
    assert "ABCDE1234F" not in clean
    assert "XXXXXX3210" in clean or "XXXX" in clean


def test_redact_csv(tmp_path: Path) -> None:
    in_csv = tmp_path / "data.csv"
    out_csv = tmp_path / "redacted_data.csv"

    with open(in_csv, "w", encoding="utf-8", newline="") as f:
        writer = csv.writer(f)
        writer.writerow(["Name", "Aadhaar", "PAN"])
        writer.writerow(["Rahul", "999988887777", "ABCDE1234F"])

    count = redact_csv(in_csv, out_csv)
    assert count == 2
    assert out_csv.exists()

    content = out_csv.read_text(encoding="utf-8")
    assert "999988887777" not in content
    assert "ABCDE1234F" not in content
    assert "Rahul" in content


def test_redact_xlsx(tmp_path: Path) -> None:
    in_xlsx = tmp_path / "data.xlsx"
    out_xlsx = tmp_path / "redacted_data.xlsx"

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["Name", "Aadhaar", "PAN"])
    ws.append(["Priya", "999988887777", "ABCDE1234F"])
    wb.save(in_xlsx)
    wb.close()

    count = redact_xlsx(in_xlsx, out_xlsx)
    assert count == 2
    assert out_xlsx.exists()

    wb_out = openpyxl.load_workbook(out_xlsx)
    ws_out = wb_out.active
    val_aadhaar = str(ws_out.cell(row=2, column=2).value)
    val_pan = str(ws_out.cell(row=2, column=3).value)

    assert "999988887777" not in val_aadhaar
    assert "ABCDE1234F" not in val_pan
    wb_out.close()
