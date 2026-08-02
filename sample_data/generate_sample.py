"""Generate a synthetic messy spreadsheet for end-to-end CLI testing.

Run: python sample_data/generate_sample.py
Creates: sample_data/messy_data.xlsx  and  sample_data/contacts.csv

v2: Now includes realistic noise patterns:
  - Multi-value cells (comma-separated identifiers)
  - Dot/underscore/slash-separated numbers
  - Labeled identifiers ("PAN: ABCPD1234E")
  - Partially masked identifiers ("XXXX XXXX 1234")
  - Excel-style numeric cells (Aadhaar stored as number, not text)
  - Currency amounts, timestamps, and invoice IDs (should NOT match)
  - Mixed-case PAN
  - Various mobile number formats

All identifiers are fabricated — none are real.
"""
from __future__ import annotations

import csv
import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).parent.parent / "src"))

import openpyxl
from phi_scanner.recognizers.aadhaar import verhoeff_check_digit
from phi_scanner.recognizers.gstin import _gstin_check_digit


def make_aadhaar(prefix: str) -> str:
    return prefix + verhoeff_check_digit(prefix)


def make_gstin(prefix14: str) -> str:
    return prefix14 + _gstin_check_digit(prefix14)


# ---------------------------------------------------------------------------
# Employee data — clean structured rows with various identifier formats
# ---------------------------------------------------------------------------

PEOPLE = [
    {
        "employee_id": "EMP001",
        "name": "Priya Sharma",
        "aadhaar": make_aadhaar("23456789012"),
        "pan": "ABCPD1234E",
        "gstin": make_gstin("29ABCPD1234E1Z"),
        "mobile": "9876543210",
        "department": "Finance",
    },
    {
        "employee_id": "EMP002",
        "name": "Rajesh Kumar",
        "aadhaar": make_aadhaar("34567890123"),
        "pan": "XYZCA5678F",
        "gstin": "",
        "mobile": "+91 8765432109",
        "department": "Engineering",
    },
    {
        "employee_id": "EMP003",
        "name": "Aisha Begum",
        # Dot-separated Aadhaar (some HR systems export this way)
        "aadhaar": ".".join([
            make_aadhaar("45678901234")[:4],
            make_aadhaar("45678901234")[4:8],
            make_aadhaar("45678901234")[8:],
        ]),
        "pan": "DEFHG2345K",
        "gstin": make_gstin("27DEFHG2345K1Z"),
        "mobile": "7654321098",
        "department": "HR",
    },
    {
        "employee_id": "EMP004",
        "name": "Suresh Pillai",
        "aadhaar": "",
        # Mixed-case PAN (copy-paste error)
        "pan": "Mnofs6789L",
        "gstin": "",
        "mobile": "6543210987",
        "department": "Sales",
    },
    {
        "employee_id": "EMP005",
        "name": "Kavita Nair",
        "aadhaar": make_aadhaar("56789012345"),
        "pan": "",
        "gstin": make_gstin("33MNOFS6789L1Z"),
        "mobile": "9123456789",
        "department": "Legal",
    },
]

# ---------------------------------------------------------------------------
# Messy data — real-world edge cases
# ---------------------------------------------------------------------------

MESSY_ROWS = [
    # Multi-value cell
    {
        "notes": f"PAN: ABCPD1234E, Mobile: 9876543210",
        "category": "combo",
    },
    # Labeled Aadhaar with misspelling
    {
        "notes": f"Aadhar No: {make_aadhaar('67890123456')}",
        "category": "labeled",
    },
    # Partially masked Aadhaar (still PII!)
    {
        "notes": "XXXX XXXX 1234",
        "category": "masked",
    },
    # Partially masked with stars
    {
        "notes": "**** **** 5678",
        "category": "masked_stars",
    },
    # Spaced PAN
    {
        "notes": "ABC PD 1234 E",
        "category": "spaced_pan",
    },
    # Slash-separated Aadhaar
    {
        "notes": f"{make_aadhaar('78901234567')[:4]}/{make_aadhaar('78901234567')[4:8]}/{make_aadhaar('78901234567')[8:]}",
        "category": "slashed_aadhaar",
    },
    # Mobile with dots
    {
        "notes": "98765.43210",
        "category": "dotted_mobile",
    },
    # Mobile with parens
    {
        "notes": "(91) 9876543210",
        "category": "parens_mobile",
    },
    # PAN in email (should NOT be flagged as PAN)
    {
        "notes": "ABCPD1234E@company.com",
        "category": "email_fp",
    },
]

# ---------------------------------------------------------------------------
# Noise rows — should NOT produce HIGH-confidence findings
# ---------------------------------------------------------------------------

NOISE_ROWS = [
    {"data": "Ref: 1234567890123 (13 digits — not Aadhaar)", "type": "order_ref"},
    {"data": "Turnover: 5000000000 (10 digit amount)", "type": "amount"},
    {"data": "20261101120000", "type": "timestamp"},
    {"data": "₹9876543210", "type": "currency"},
    {"data": "INR 8765432109", "type": "inr_amount"},
    {"data": "Rs. 7654321098", "type": "rs_amount"},
    {"data": "INV-9876543210", "type": "invoice"},
    {"data": "ORD#8765432109", "type": "order_id"},
    {"data": "EMP5432109876", "type": "emp_id"},
]


def write_xlsx(out: Path) -> None:
    wb = openpyxl.Workbook()

    # Sheet 1: Clean employee data
    ws_emp = wb.active
    ws_emp.title = "Employees"
    headers = list(PEOPLE[0].keys())
    ws_emp.append(headers)
    for p in PEOPLE:
        ws_emp.append([p[h] for h in headers])

    # Sheet 2: Messy data
    ws_messy = wb.create_sheet("Messy")
    ws_messy.append(["notes", "category"])
    for row in MESSY_ROWS:
        ws_messy.append([row["notes"], row["category"]])

    # Sheet 3: Aadhaar stored as numbers (Excel numeric cells)
    ws_numeric = wb.create_sheet("Numeric_IDs")
    ws_numeric.append(["aadhaar_no", "phone_no", "name"])
    valid_aadhaar = int(make_aadhaar("89012345678"))
    ws_numeric.append([valid_aadhaar, 9876543210, "Test Person"])
    ws_numeric.append([int(make_aadhaar("90123456789")), 8765432109, "Another Person"])

    # Sheet 4: Noise (should not match)
    ws_noise = wb.create_sheet("Noise")
    ws_noise.append(["data", "type"])
    for n in NOISE_ROWS:
        ws_noise.append([n["data"], n["type"]])

    wb.save(out)
    print(f"Written: {out}")


def write_csv(out: Path) -> None:
    fieldnames = ["name", "mobile", "aadhaar_no", "pan_no"]
    rows = [
        {"name": p["name"], "mobile": p["mobile"],
         "aadhaar_no": p["aadhaar"], "pan_no": p["pan"]}
        for p in PEOPLE
    ]
    # Add a multi-value row
    rows.append({
        "name": "Multi-value cell",
        "mobile": "9876543210, 8765432109",
        "aadhaar_no": f"UID: {make_aadhaar('23456789012')}",
        "pan_no": "PAN: ABCPD1234E",
    })
    with open(out, "w", newline="", encoding="utf-8") as fh:
        writer = csv.DictWriter(fh, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(rows)
    print(f"Written: {out}")


if __name__ == "__main__":
    sample_dir = Path(__file__).parent
    write_xlsx(sample_dir / "messy_data.xlsx")
    write_csv(sample_dir / "contacts.csv")
